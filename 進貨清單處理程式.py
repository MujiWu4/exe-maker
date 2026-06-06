# ==========================================
# [第一部分] 升級版：只有當「數量」與「庫存總量」完全相同時，才標記綠色！
# ==========================================
import os
import re
from tkinter import Button, Label, Tk, filedialog, messagebox
from openpyxl.styles import PatternFill
import pandas as pd


def extract_keywords(name):
    """【名偵探的放大鏡】把品名『脫掉外套』，只留下最核心的品牌和型號！"""
    if pd.isna(name):
        return []
    name_clean = str(name).upper()
    name_clean = re.sub(r"[#（）()\[\]【】\-/\s]", " ", name_clean)
    tokens = name_clean.split()
    ignore_words = {"個", "把", "台", "支", "組", "副", "張", "包", "套", "條"}
    keywords = [
        t for t in tokens if len(t) > 1 and t not in ignore_words and not t.isdigit()
    ]
    if not keywords:
        keywords = [t for t in tokens if t not in ignore_words]
    return keywords


def match_stock(purchase_row, df_stock):
    """【連連看大考驗】拿著進貨的『貨品名稱』去比對庫存，找出最高分商品！"""
    p_name = purchase_row.get("貨品名稱", "")
    if pd.isna(p_name) or str(p_name).strip() == "":
        return None
    keywords = extract_keywords(p_name)
    if not keywords:
        return None
    best_match_idx = None
    max_score = 0
    for idx, stock_row in df_stock.iterrows():
        stock_name = (
            str(stock_row["商品名稱"]).upper() + " " + str(stock_row["商品款式"]).upper()
        )
        score = sum(1 for kw in keywords if kw in stock_name)
        if score > max_score:
            max_score = score
            best_match_idx = idx
    if max_score > 0:
        return df_stock.loc[best_match_idx]
    return None


def scan_and_merge_purchase_all_columns(file_path):
    """完整保留進貨單所有出現過的欄位資料，並進行跨段落合併"""
    df_raw = pd.read_excel(file_path, header=None)
    all_data_rows, current_columns, is_collecting = [], None, False
    req_cols = ["貨品編號", "貨品名稱"]
    for idx, row in df_raw.iterrows():
        row_values = [
            str(val).strip() if pd.notna(val) else "" for val in row.values
        ]
        if all(col in row_values for col in req_cols):
            temp_cols = []
            for i, v in enumerate(row_values):
                temp_cols.append(v if v != "" else f"__EMPTY_COL_{i}__")
            current_columns = temp_cols
            is_collecting = True
            continue
        if is_collecting:
            if "進貨日期" in "".join(row_values) or "進貨清單" in "".join(row_values):
                is_collecting = False
                continue
            if "".join(row_values) == "":
                is_collecting = False
                continue
            row_dict = {}
            for col_name, val in zip(current_columns, row.values):
                if "__EMPTY_COL_" in col_name:
                    if pd.notna(val) and str(val).strip() != "":
                        row_dict[col_name] = val
                else:
                    row_dict[col_name] = val if pd.notna(val) else ""
            if not row_dict.get("貨品編號") and not row_dict.get("貨品名稱"):
                if not row_dict.get("貨品名稱"):
                    continue
            all_data_rows.append(row_dict)
    if not all_data_rows:
        return None
    df_merged = pd.DataFrame(all_data_rows)
    cleaned_columns = [
        "" if "__EMPTY_COL_" in str(col) else col for col in df_merged.columns
    ]
    df_merged.columns = cleaned_columns
    return df_merged


def smart_load_stock(file_path):
    """自動尋找庫存清單標題列"""
    df_raw = pd.read_excel(file_path, header=None)
    required_cols = ["商品名稱", "庫存總量"]
    for idx, row in df_raw.iterrows():
        row_values = [str(val).strip() for val in row.values if pd.notna(val)]
        if all(col in row_values for col in required_cols):
            df_correct = pd.read_excel(file_path, header=idx)
            df_correct.columns = df_correct.columns.str.strip()
            return df_correct
    return None


def start_matching(purchase_path, stock_path):
    if not purchase_path or not stock_path:
        messagebox.showwarning("提示", "請先選擇【進貨清單】和【庫存清單】檔案！")
        return
    try:
        df_purchase = scan_and_merge_purchase_all_columns(purchase_path)
        df_stock = smart_load_stock(stock_path)
        if df_purchase is None or df_stock is None:
            messagebox.showerror("錯誤", "找不到對應欄位，請檢查檔案名稱！")
            return
        stock_names_col, stock_styles_col, stock_totals_col = [], [], []
        for _, p_row in df_purchase.iterrows():
            matched_stock_row = match_stock(p_row, df_stock)
            if matched_stock_row is not None:
                stock_names_col.append(matched_stock_row["商品名稱"])
                stock_styles_col.append(matched_stock_row.get("商品款式", ""))
                # 確保庫存數量轉成數字，抓不到就當作 0
                try:
                    stock_totals_col.append(int(matched_stock_row["庫存總量"]))
                except:
                    stock_totals_col.append(0)
            else:
                p_name = str(p_row.get("貨品名稱", "")).strip()
                if p_name == "" or p_name == "nan":
                    stock_names_col.append("")
                    stock_styles_col.append("")
                    stock_totals_col.append("")
                else:
                    stock_names_col.append("⚠️ 庫存系統找不到這個商品")
                    stock_styles_col.append("")
                    stock_totals_col.append(0)
        df_purchase["商品名稱_庫存系統"] = stock_names_col
        df_purchase["商品款式_庫存系統"] = stock_styles_col
        df_purchase["庫存總量"] = stock_totals_col

        current_cols = list(df_purchase.columns)
        cols_to_move = ["庫存總量"]
        cols_to_end = ["商品名稱_庫存系統", "商品款式_庫存系統"]
        for col in cols_to_move + cols_to_end:
            if col in current_cols:
                current_cols.remove(col)
        if "單位" in current_cols:
            idx_unit = current_cols.index("單位")
            new_col_order = (
                current_cols[: idx_unit + 1]
                + cols_to_move
                + current_cols[idx_unit + 1 :]
                + cols_to_end
            )
        else:
            new_col_order = current_cols + cols_to_move + cols_to_end
        df_output = df_purchase[new_col_order]

        output_dir = os.path.dirname(purchase_path)
        output_filename = os.path.join(output_dir, "完整進貨單_對應庫存總表.xlsx")
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            df_output.to_excel(writer, index=False, sheet_name="對帳結果")
            workbook = writer.book
            worksheet = writer.sheets["對帳結果"]

            # 舒服的淡粉綠色螢光筆
            green_fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )

            # 尋找目標直行（欄位位置）
            qty_idx = (
                new_col_order.index("數量") + 1 if "數量" in new_col_order else None
            )
            stock_total_idx = (
                new_col_order.index("庫存總量") + 1
                if "庫存總量" in new_col_order
                else None
            )
            check_name_idx = new_col_order.index("商品名稱_庫存系統") + 1

            # 螢光筆小精靈出動巡邏
            for row_idx in range(2, worksheet.max_row + 1):
                check_value = worksheet.cell(
                    row=row_idx, column=check_name_idx
                ).value

                # 1. 先確認是不是有效的庫存商品
                if (
                    check_value
                    and check_value != "⚠️ 庫存系統找不到這個商品"
                    and str(check_value).strip() != ""
                ):

                    # 2. 🔥【核心魔法】抓出這兩格的數值，並轉換成整數來比較
                    raw_qty = (
                        worksheet.cell(row=row_idx, column=qty_idx).value
                        if qty_idx
                        else 0
                    )
                    raw_stock = (
                        worksheet.cell(row=row_idx, column=stock_total_idx).value
                        if stock_total_idx
                        else 0
                    )

                    try:
                        qty_val = (
                            int(float(raw_qty))
                            if pd.notna(raw_qty) and raw_qty != ""
                            else 0
                        )
                        stock_val = (
                            int(float(raw_stock))
                            if pd.notna(raw_stock) and raw_stock != ""
                            else 0
                        )
                    except ValueError:
                        qty_val, stock_val = (
                            -1,
                            -2,
                        )  # 如果有字串沒辦法轉數字，就讓它們不相等

                    # 3. 只有當【進貨數量】等於【庫存總量】時，魔法螢光筆才畫下去！
                    if qty_val == stock_val:
                        if qty_idx:
                            worksheet.cell(
                                row=row_idx, column=qty_idx
                            ).fill = green_fill
                        if stock_total_idx:
                            worksheet.cell(
                                row=row_idx, column=stock_total_idx
                            ).fill = green_fill

        messagebox.showinfo(
            "成功",
            f"🎉 條件對帳著色完成！\n只有當【數量】與【庫存總量】相同時才標記綠色！\n檔案已儲存在：\n{output_filename}",
        )
    except Exception as e:
        messagebox.showerror("錯誤", f"發生錯誤：{str(e)}")
# ==========================================
# [第二部分] 這是程式的五官：也就是你貼上來的那段視窗控制與啟動按鈕
# ==========================================
class MatcherGUI:

    def __init__(self, root):
        self.root = root
        self.root.title("樂器行 AI 庫存對帳小幫手 v1.5 (全自動綠色螢光筆版)")
        self.root.geometry("550x300")

        self.purchase_file_path = ""
        self.stock_file_path = ""

        self.title_label = Label(
            root, text="歡迎來到資優班庫存比對系統", font=("微軟正黑體", 14, "bold")
        )
        self.title_label.pack(pady=15)

        self.btn_purchase = Button(
            root,
            text="1. 選擇【進貨清單】Excel (保留全欄位)",
            command=self.select_purchase,
            width=32,
        )
        self.btn_purchase.pack(pady=5)

        self.lbl_purchase = Label(
            root, text="尚未選擇檔案...", fg="gray", wraplength=450
        )
        self.lbl_purchase.pack(pady=2)

        self.btn_stock = Button(
            root,
            text="2. 選擇【庫存清單】Excel",
            command=self.select_stock,
            width=32,
        )
        self.btn_stock.pack(pady=5)

        self.lbl_stock = Label(root, text="尚未選擇檔案...", fg="gray", wraplength=450)
        self.lbl_stock.pack(pady=2)

        self.btn_run = Button(
            root,
            text="🚀 啟動全智慧視覺化對帳",
            command=lambda: start_matching(
                self.purchase_file_path, self.stock_file_path
            ),
            bg="#2E7D32",
            fg="white",
            font=("微軟正黑體", 11, "bold"),
            width=30,
        )
        self.btn_run.pack(pady=20)

    def select_purchase(self):
        file_path = filedialog.askopenfilename(
            title="選擇進貨清單", filetypes=[("Excel 檔案", "*.xlsx *.xls")]
        )
        if file_path:
            self.purchase_file_path = file_path
            filename = os.path.basename(file_path)
            self.lbl_purchase.config(text=f"已選取：{filename}", fg="green")

    def select_stock(self):
        file_path = filedialog.askopenfilename(
            title="選擇庫存清單", filetypes=[("Excel 檔案", "*.xlsx *.xls")]
        )
        if file_path:
            self.stock_file_path = file_path
            filename = os.path.basename(file_path)
            self.lbl_stock.config(text=f"已選取：{filename}", fg="green")


# 啟動魔法視窗的核心心臟
if __name__ == "__main__":
    window = Tk()
    app = MatcherGUI(window)
    window.mainloop()
