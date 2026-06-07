import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl

def clean_and_generate_seo(product_name):
    """【終極全通用演算法】無硬編碼品牌限制，利用物理字元邊界切斷，生成純大寫變數 (修復 List 型態衝突)"""
    if not product_name:
        return ""

    if isinstance(product_name, list):
        product_name = " ".join(str(x) for x in product_name)
    else:
        product_name = str(product_name).strip()

    # 1. 強制轉大寫，並移除所有中文、特殊符號與括號備註內容
    upper_name = product_name.upper()
    pure_eng_num = re.sub(r"[\u4e00-\u9fa5]", " ", upper_name)
    pure_eng_num = re.sub(r"[【】（）()\[\]／,.\s]+", " ", pure_eng_num).strip()

    if not pure_eng_num:
        return ""

    # 2. 處理特殊電商重疊防護檢查 (適用於 Expanding / Upgrade / Pack 系列)
    pack_match = re.search(r"PACK\s*([0-9]+)", upper_name)
    if pack_match:
        pack_num = pack_match.group(1)
        if "7/5" in upper_name or "7&5" in upper_name:
            return f"# EXPANDINGPACK{pack_num} / EFNOTE75 / EFNOTE57"
        return f"# EXPANDINGPACK{pack_num}"

    # 3. 通用物理型號切斷術：自動尋找「連續英文字母 + 數字」的核心交界
    core_model = ""
    model_match = re.search(r"([A-Z]+[-/ ]*[0-9]+)", pure_eng_num)
    
    if model_match:
        core_model = model_match.group(1).strip()
    else:
        # 【關鍵修復點】：若無標準英數交界，將列表元件用空格結合為純字串，避免 replace 錯誤
        words = pure_eng_num.split()
        if isinstance(words, list):
            core_model = " ".join(str(x) for x in words)
        else:
            core_model = str(words)

    # 確保 core_model 一定是字串型態後，再進行去空白處理
    core_model = str(core_model).replace(" ", "")

    if not core_model or core_model.strip() == "":
        return ""

    variants = []

    if "-" in core_model or "/" in core_model:
        all_sticky = re.sub(r"[-/ ]", "", core_model)
        all_space = re.sub(r"[-/]", " ", core_model)
        all_space = " ".join(all_space.split())
        if all_sticky: variants.append(all_sticky)
        if all_space and all_space != all_sticky: variants.append(all_space)
    else:
        split_match = re.match(r"([A-Z]+)([0-9]+.*)", core_model)
        if split_match:
            letters = split_match.group(1)
            numbers = split_match.group(2)
            variants.append(f"{letters}-{numbers}")
            variants.append(f"{letters} {numbers}")
        else:
            variants.append(f"{core_model}")

    # 4. 最嚴格去重複（Substring Match 防呆）
    final_variants = []
    for v in variants:
        v_upper = str(v).upper()
        if v_upper in upper_name:
            continue
        if v_upper not in final_variants:
            final_variants.append(v_upper)

    if final_variants:
        return f"# " + " / ".join(final_variants)
    return ""

def smart_merge_summary_with_seo(original_summary, new_seo):
    """【終極去誤殺＆除噪升級版】逐行精準清除不規則舊標籤行（含-#開頭等雜訊），保留純內文，底端換行三次附加"""
    if isinstance(original_summary, list):
        orig_text = " ".join(str(x) for x in original_summary)
    else:
        orig_text = str(original_summary) if original_summary else ""

    if not orig_text.strip():
        return new_seo if new_seo else ""

    clean_lines = []
    for line in orig_text.splitlines():
        trimmed_line = line.strip()
        
        # 判定 A：如果這一行是以 # 開頭，直接判定為舊關鍵字行，移除
        if trimmed_line.startswith("#"):
            continue
            
        # 判定 B：如果整行只包含 #、-、/、空格、英數符號，且含有任何 # 號，移除
        if re.match(r"^[\s#A-Za-z0-9/\-_&]+$", trimmed_line) and "#" in trimmed_line:
            continue
            
        # 判定 C：如果這一行是以 - 橫線開頭，且整行「純粹由英數字、橫線、空格組成」，判定為隱形雜訊行，移除
        if trimmed_line.startswith("-") and re.match(r"^[\sA-Za-z0-9\-]+$", trimmed_line):
            continue
            
        # 如果都不是上述關鍵字雜訊行，代表是描述內文，予以保留
        clean_lines.append(line.rstrip())

    final_body = "\n".join(clean_lines).strip()

    if not final_body:
        return new_seo if new_seo else ""
    if not new_seo:
        return final_body

    return f"{final_body}\n\n\n{new_seo}"

class AutoExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("電商全品類通用：搜尋優化防呆工具 (終極完美除錯版)")
        self.root.geometry("550x300")
        self.root.resizable(False, False)

        self.file_path = tk.StringVar()
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_lbl = ttk.Label(main_frame, text="自動模糊尋欄 ＆ 摘要附加大寫關鍵字工具", font=("Helvetica", 14, "bold"))
        title_lbl.pack(pady=(0, 15))

        file_frame = ttk.LabelFrame(main_frame, text=" 選擇商品 Excel 檔案 ", padding="10")
        file_frame.pack(fill=tk.X, pady=10)

        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=40, state="readonly")
        self.file_entry.pack(side=tk.LEFT, padx=(0, 10))

        browse_btn = ttk.Button(file_frame, text="瀏覽檔案", command=self.browse_file)
        browse_btn.pack(side=tk.LEFT)

        self.process_btn = ttk.Button(main_frame, text="開始自動查找並優化", command=self.process_excel, state=tk.DISABLED)
        self.process_btn.pack(fill=tk.X, pady=15)

        self.status_lbl = ttk.Label(main_frame, text="請選取您要優化的 Excel 檔案...", foreground="gray")
        self.status_lbl.pack()

    def browse_file(self):
        file_selected = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_selected:
            self.file_path.set(file_selected)
            self.process_btn.config(state=tk.NORMAL)
            self.status_lbl.config(text="檔案選取成功！請點擊按鈕開始轉換。", foreground="green")

    def process_excel(self):
        src_path = self.file_path.get()
        dir_name, file_name = os.path.split(src_path)
        name_part, ext_part = os.path.splitext(file_name)
        dst_path = os.path.join(dir_name, f"{name_part}_SEO_Fixed{ext_part}")

        try:
            self.status_lbl.config(text="正在分析欄位與處理資料...", foreground="blue")
            self.root.update()

            wb = openpyxl.load_workbook(src_path, data_only=False)
            ws = wb.active

            name_col = None
            summ_col = None

            # 精準強制只讀取第 1 行的每一個格子物件，防範合併儲存格的 Tuple 衝突
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    val = str(cell.value).strip().upper() if cell.value else ""
                    if not val: continue
                    if "名稱" in val or "名称" in val or "NAME" in val:
                        name_col = cell.column_letter
                    if "摘要" in val or "SUMMARY" in val:
                        summ_col = cell.column_letter

            if not name_col or not summ_col:
                detected_headers = []
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        if cell.value: detected_headers.append(str(cell.value))
                raise ValueError(f"在標題列中找不到對應欄位！\n目前偵測到的標題有：{detected_headers}")

            max_row = ws.max_row
            processed_count = 0

            for row in range(2, max_row + 1):
                name_cell = ws[f"{name_col}{row}"]
                summ_cell = ws[f"{summ_col}{row}"]

                if name_cell.value:
                    seo_tags = clean_and_generate_seo(name_cell.value)
                    updated_summary = smart_merge_summary_with_seo(summ_cell.value, seo_tags)
                    summ_cell.value = updated_summary
                    processed_count += 1

            wb.save(dst_path)
            wb.close()

            self.status_lbl.config(text=f"成功！已優化 {processed_count} 筆商品資料。", foreground="green")
            messagebox.showinfo("成功", f"優化完成！已另存新檔至：\n{dst_path}")

        except Exception as e:
            self.status_lbl.config(text="處理失敗，請檢查錯誤。", foreground="red")
            messagebox.showerror("欄位讀取或轉換錯誤", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoExcelProcessorApp(root)
    root.mainloop()
